from tkinter import *
import os               
import re
import pandas as pd
import jinja2
import openpyxl
from openpyxl.styles import Font

from tkinter import filedialog as fd

class Timeline():
    def __init__(self, root):
        self.master = root
        self.master.title("Timeline Input")
# labels
        self.lblDate = Label(self.master, text="Date")
        self.lblDate.grid(row=0,column=0, sticky=W, padx=5)
        self.lblLoc = Label(self.master, text="Loc")
        self.lblLoc.grid(row=0,column=1, sticky=W, padx=5)
        self.lblCoord = Label(self.master, text="Coordinates")
        self.lblCoord.grid(row=0,column=2, sticky=W, padx=5)
        self.lblLink = Label(self.master, text="Web Link")
        self.lblLink.grid(row=0,column=3, sticky=W, padx=5)

# entryboxes
        self.eboxDate = Text(self.master,width=20,height=1, font=('Chirp 14'))
        self.eboxDate.grid(row=1, column=0, padx=5)
        self.eboxLoc = Text(self.master,width=20,height=1, font=('Chirp 14'))
        self.eboxLoc.grid(row=1, column=1, padx=5)
        self.eboxCoord = Text(self.master,width=20,height=1, font=('Chirp 14'))
        self.eboxCoord.grid(row=1, column=2, padx=5)
        self.eboxLink = Text(self.master,width=30,height=1, font=('Chirp 14'))
        self.eboxLink.grid(row=1, column=3, padx=5)

# buttons
        self.btnQuit = Button (self.master, text="Quit", command=self.bQuit)
        self.btnQuit.grid(row=2,column=0, sticky=W, padx=5)
        self.btnSubmit = Button (self.master, text="Submit", command=self.bSubmit)
        self.btnSubmit.grid(row=2,column=3, sticky=E, padx=5)

    def bSubmit(self):
        df = pd.DataFrame([[self.eboxDate.get('1.0',END), self.eboxLoc.get('1.0',END), self.eboxCoord.get('1.0',END), self.eboxLink.get('1.0',END)]], columns=['Date','Location','Coord','Link'])
        df = df.style.set_properties(**{'font-size': '14pt'})

        if(os.path.exists('Timeline.xlsx') == False):                                  # check for excel doc existing
            print('file doesnt exist - creating')
            writer = pd.ExcelWriter('Timeline.xlsx', engine='xlsxwriter')              # create excel doc 
            df.to_excel(writer, index=False)
            writer.save()

            self.sheetlayout()

        else:
            print('file exists')
            if(pd.read_excel('Timeline.xlsx').empty):
                print('- file is empty')                                                # check is excel doc is empty
                print(df)
                
                self.sheetlayout()

            else:
                print("not empty")
                wb = openpyxl.load_workbook('Timeline.xlsx')
                ws = wb.active
                ws.insert_rows(1)
                print(ws.max_row)

                ws.append([self.eboxDate.get('1.0',END), self.eboxLoc.get('1.0',END), self.eboxCoord.get('1.0',END), self.eboxLink.get('1.0',END)])

                rowindex = '{}:{}'.format(ws.max_row, ws.max_column)
                for cell in ws[rowindex]:
                    cell.font = Font(name='Chirp', size=14)

                self.sheetlayout()
                wb.save('Timeline.xlsx')

    def bQuit(self):
        TkRoot.destroy()

    def sheetlayout(self):
        print('setting column spacing')
        # set visual properties
        ws = openpyxl.load_workbook('Timeline.xlsx')
        sheet = ws.active
        sheet.column_dimensions['C'].width = 8.43 * 4
        sheet.column_dimensions['B'].width = 8.43 * 4       # width of columns as strings vary in length
        sheet.column_dimensions['C'].width = 8.43 * 6
        sheet.column_dimensions['D'].width = 8.43 * 6
        ws.save('Timeline.xlsx')

if __name__ == '__main__':
    TkRoot = Tk()
    my_gui = Timeline(TkRoot)

    TkRoot.mainloop()